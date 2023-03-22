import os #pip install os-sys
import os.path #pip install os.path2
import zipfile #pip install zipfile36
import shutil #pip install pytest-shutil
import re #pip install regex
import fnmatch
import pandas as pd #pip install pandas
import openpyxl #pip install openpyxl
import PyPDF2 #pip install PyPDF2
import tabula #pip install tabula-py
import itertools #pip install more-itertools
import xlsxwriter
from zipfile import ZipFile
from PyPDF2 import PdfReader
from datetime import datetime #pip install datetime
from openpyxl import Workbook, load_workbook

#Funções:
def horas(number):
    return str(number)[:2]+':'+str(number)[2:]

#-------Descompactar o Zip-------
arq = os.listdir('C:\VS Code\ELW_01') #Diretório atual
arq_pdf = os.listdir('C:\VS Code\ELW_01\PDF') #Diretório do PDF
extensao = ".zip" #Definindo a extensão do arquivo
nova_pasta = "C:\VS Code\ELW_01\PDF" #Caminho da pasta PDF

for file in arq: #Descompactando os arquivos de dentro do zip e os enviando para uma pasta
    if zipfile.is_zipfile(file):
        with zipfile.ZipFile(file) as item: 
           item.extractall('C:\VS Code\ELW_01\PDF') 

for i in arq: #Excluindo o arquivo .zip
    if i.endswith(extensao):
        file_name = os.path.abspath(i)
        os.remove(file_name)

#-----Manipulando o Samples-------
samples_path = "C:\VS Code\ELW_01\Samples.pdf" #Definindo o caminho
pdf_samples = PdfReader(open(samples_path,'rb')) #Lendo o PDF
tabula_file = tabula.convert_into(samples_path, "Samples.csv", output_format="csv", pages="all") #Abrindo com o tabula
df = pd.read_csv("C:\VS Code\ELW_01\Samples.csv", encoding = "ISO-8859-1", on_bad_lines='skip') #Convertendo o Samples em csv
df.head() #Retornando o pandas

#-------Criando o Excel------
excel_directory = os.chdir("C:\VS Code\ELW_01\Excel") #Caminho da pasta do excel
excel_arq = os.listdir("C:\VS Code\ELW_01\Excel") #Pegando os arquivos da lista
if excel_arq == []: #Criando o Excel
    arquivo = xlsxwriter.Workbook("ELISA.xlsx") #Criando e nomenando o arquivo
    planilha = arquivo.add_worksheet("Consolidado_ELW 1 e 2 ELISA") #Nomeando a planilha
    cell_format = arquivo.add_format({'bg_color': 'gray', 'font_color': 'white', 'align':'center'}) #Formatação da celula
    planilha.set_column('A1:Q1', 25) #Formatação da coluna
    #Cabeçalho
    planilha.write('A1', 'Máquina', cell_format)
    planilha.write('B1', 'Nome da rotina', cell_format)
    planilha.write('C1', 'QTD Placas', cell_format)
    planilha.write('D1', 'Data de início', cell_format)
    planilha.write('E1', 'Data de término', cell_format)
    planilha.write('F1', 'Hora de início Cadastro da Rotina', cell_format)
    planilha.write('G1', 'Hora de Início (Start da Rotina)', cell_format)
    planilha.write('H1', 'Hora do término', cell_format)
    planilha.write('I1', 'Dia da Semana', cell_format)
    planilha.write('J1', 'Turno', cell_format)
    planilha.write('K1', 'QTD Resultados Liberados', cell_format)
    planilha.write('L1', 'Qtd Amostras', cell_format)
    planilha.write('M1', 'Placas processadas', cell_format)
    planilha.write('N1', 'Tempo de Carregamento da Rotina', cell_format)
    planilha.write('O1', 'Tempo Processamento Placas', cell_format)
    planilha.write('P1', 'Tempo Total Rotina (Montagem + Execução)', cell_format)
    planilha.write('Q1', 'Intervalo entre rotinas', cell_format)
    planilha.autofilter('A1:Q1')
    arquivo.close() #Fechando o Excel

#------Manipulando o Target--------
target_final = "Target_Layout.pdf" #Definindo a variavel
for i in arq_pdf: #Criando o loop para procurar o Target
    if i.endswith(target_final):
        target_name = nova_pasta + "\\" + i #Definindo o caminho de todos os arquivos target
        target_file = PdfReader(open(target_name,'rb')) #Ler o PDF do Target
        #Transformando o Target em texto
        primeira_pagina = target_file.pages[0] #Transformando a primeira página
        primeira_content = primeira_pagina.extract_text() #Extrai apenas o texto
        target1 = ''.join(primeira_content) #Faz a junção das linhas
        #Transformando o Target inteiro em texto
        target_content = [] #Criando a lista vazia
        count = len(target_file.pages) #Contando as páginas do Taget
        for i in range(count):
            target_page = target_file.pages[i] #Transformando o arquivo
            target_content += target_page.extract_text() #extrai apenas o texto
            texto_target = ''.join(target_content) #Faz a junção das linhas 

            #-------Nome da máquina-------
            nome = "ELW 1"
            print("Nome:", nome)
            #------Nome da Rotina-------
            worklist = (re.findall(r"Worklist: \s+(\w+)", target1)).pop(0) #Pegando a worklist
            print("Worklist:", worklist)
            #-----Quantidade de Placas-------
            ID = len(re.findall(r"ID: \d{2}", texto_target)) #Pegando a QTD de placas
            print("QTD Placas:", ID)
            #------Data de inicio------
            data = (re.findall(r'Date: \s+(\d+\/\d+\/\d+)', target1)).pop(0) #Pegando a data de inicio
            print("Data de inicio:", data)

            #------Manipulando o Result Report-------
            rr_list = os.listdir(nova_pasta)
            for r in rr_list: #Criando o loop para achar o Result Report
                if r.startswith(worklist) and fnmatch.fnmatch(r, '*_[0-9][0-9]_*'): #Criando a pesquisa do Result Report
                    rr_name = nova_pasta + "\\" + r #Criando a variavel que guarda o caminho do Result Report
                    resultreport_file = PdfReader(open(rr_name, 'rb')) #Lendo o Result Report
                    #----- Transformando os Result Report em texto
                    report_page = resultreport_file.pages[0] #Primeira página - Result Report
                    report_content = report_page.extract_text() #Extrai apenas o texto
                    texto_report = ''.join(report_content) #Faz a junção das linhas 
                    #-----Pegando a Data de termino-----
                    data_finish = (re.findall(r'\s+(\d+\/\d+\/\d+)', texto_report)).pop(0) #Pegando a hora de termino
                    print(data_finish)
                    #--------Pegando a Hora de termino-----
                    hora_finish = ((re.findall(r'\s+(\d+\:\d+\s+\w+)', texto_report))).pop(0) #Pegando a hora de termino
                    hora_finish = datetime.strptime(hora_finish, "%I:%M %p") #Fazendo a formatação (transformando para time)
                    hora_finish = datetime.strftime(hora_finish, "%H:%M") #Fazendo a formatação (transformando para string)
                    print("Hora de termino:", hora_finish)
            
            #-----Pegando a Hora de inicio-----
            hora_inicio = (re.findall(r'_\d{4}', target1)).pop(0) #Pegando a hora de inicio
            hora_inicio = (re.findall(r'\d{4}', hora_inicio)).pop(0) #Pegando a hora de inicio e tirando a _
            hora_inicio = horas(hora_inicio)
            print("Hora de inicio:", hora_inicio)

            #------Pegando a Hora de inicio (Start da rotina)----
            hora_start = ((re.findall(r'\s+(\d+\:\d+\s+\w+)', target1))).pop(0) #Pegando a hora de start
            hora_start = datetime.strptime(hora_start, "%I:%M %p") #Fazendo a formatação (transformando para time)
            hora_start = datetime.strftime(hora_start, "%H:%M") #Fazendo a formatação (transformando para string)
            print("Hora de inicio (Start da rotina):", hora_start)

            #-----Pegando o Dia da semana----
            dds = datetime.strptime(data, '%m/%d/%Y') #Fazendo a formatação (transformando para time)
            dds = datetime.strftime(dds, "%A") #Fazendo a formatação (transformando para string)
            if dds == "Monday": #Criando a condicional para retornar o valor do dia
                dds = "Seg"
            elif dds == "Tuesday":
                dds = "Ter"
            elif dds == "Wednesday":
                dds = "Qua"
            elif dds == "Thursday":
                dds = "Qui"
            elif dds == "Friday":
                dds = "Sex"
            elif dds == "Saturday":
                dds = "Sáb"
            elif dds == "Sunday":
                dds = "Dom"
            print("Dia da semana:", dds)

            #----Pegando o turno----
            t = hora_inicio
            hora1 = int(300) #Transformando 05:00 em minutos
            hora2 = int(1080) #Transformando 18:00 em minutos
            h1 = t[:2] #Pegando as strings
            h2 = t[2:] #Pegando as strings
            h2 = (re.findall(r'\d{2}', h2)).pop(0) #Tirando o :
            A = [h1] #Colocando em uma lista
            B = [h2] #Colocando em uma lista 
            res = [ele.lstrip('0') for ele in A].pop(0)
            res2 = [ele.lstrip('0') for ele in B].pop(0)
            
            h1int = int(res) #Transformando em int
            h2int = int(res2) #Transformando em int
            horas_minutos = h1int * 60 #Transformando em minutos
            minutos = int(horas_minutos + h2int) #Pegando o valor da hora de inicio em minutos
    
            d = "Diurno" 
            n = "Noturno"
            if (minutos >= hora1):
                t = "Diurno"
            elif (minutos >= hora2):
                t = "Noturno"
            else:
                t = "Noturno"
            print("Turno:", t)

            #-----Pegando QTD Resultados Liberados----
            qtd_rl = len(re.findall(r'\d{10}', texto_target)) #Buscando no target
            print("QTD Resultados Liberados", qtd_rl)

            #------Pegando Qtd Amostras------
            with open('C:\VS Code\ELW_01\Samples.csv') as f:
                qtd_a = f.read().count(worklist) #Contando a worklist no Samples
            print("QTD Amostras:", qtd_a)

            #-----Pegando Placas processadas----
            placas = str(round(qtd_rl/96, 2)).replace('.',',') #Substituindo o . por ,
            print("Placas processadas:", placas)

            #-----Pegando Tempo de Carregamento da Rotina----
            temp_start = datetime.strptime(hora_start, "%H:%M") #Convertendo para horas
            temp_start = temp_start - datetime(1900, 1, 1) #Fazendo a formatação
            temp_inicio = datetime.strptime(hora_inicio, "%H:%M") #Convertendo para horas
            temp_inicio = temp_inicio - datetime(1900, 1, 1) #Fazendo a formatação
            tcr = temp_start - temp_inicio #Fazendo a subtração
            print("Tempo de Carregamento da rotina:", tcr)

            #-------Tempo Processamento Placas-----
            data_i = datetime.strptime(data, '%m/%d/%Y') #Convertendo para horas
            hora_is = datetime.strptime(hora_start, "%H:%M") - datetime(1900, 1, 1) #Fazendo a formatação
            data_f = datetime.strptime(data_finish, '%m/%d/%Y') #Convertendo para horas 
            hora_f = datetime.strptime(hora_finish, "%H:%M") - datetime(1900, 1, 1) #Fazendo a formatação
            tpp = (data_f + hora_f) - (data_i + hora_is) #Fazendo a subtração
            print("Tempo Processamento Placas:", tpp)

            #------Tempo Total Rotina (Montagem + Execução)-----
            ttp = tcr + tpp
            print("Tempo Total Rotina (Montagem + Execução):", ttp)

            #----Manipulando o Excel-----
            elw = load_workbook("C:\VS Code\ELW_01\Excel\ELISA.xlsx")
            elw1 = elw.active
            elw1["A2"].value = "Teste"
            elw.save("ELISA.xlsx")

            print("-----------------")