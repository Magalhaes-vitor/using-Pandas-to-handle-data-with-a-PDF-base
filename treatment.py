import os
import re
import glob
import tabula
import shutil
import fnmatch
import os.path
import zipfile
import openpyxl
import itertools
import xlsxwriter
import pandas as pd
from zipfile import ZipFile
from PyPDF2 import PdfReader
from datetime import datetime
from openpyxl import Workbook, load_workbook
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
arqs = os.listdir() #Diretorio atual
extensao = ".zip" #Definimos a extensão do objeto
diretorio = r'C:\Users\vitor.magalhaes\Desktop\canudo\V1\PDF\ELW_01'
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
for file in arqs: #Descompactando os arquivos de dentro do zip e os enviando para uma pasta 
    if zipfile.is_zipfile(file): 
        with zipfile.ZipFile(file) as item: 
           item.extractall(diretorio + '\PDF')  
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
for i in arqs: #Excluindo o arquivo .zip
    if i.endswith(extensao): 
        file_name = os.path.abspath(i)
        os.remove(file_name)
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
diretorio_pdf = diretorio + '\PDF' #Enviando o Samples para o novo diretório
shutil.move( diretorio + "\Samples.pdf", diretorio_pdf) # só funciona se o arquivo ainda não movido
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
excel_directory = os.chdir(diretorio + '\Excel')#Criar o excel
excel_arq = os.listdir()
if excel_arq == []: #Criando o Excel
    arquivo = xlsxwriter.Workbook("ELW ELISA 1 -  Pardini - Produtividade 14 Dias.xlsx") #Criando e nomenando o arquivo
    planilha = arquivo.add_worksheet("Consolidado_ELW 1 e 2 ELISA") #Nomeando a planilha
    cell_format = arquivo.add_format({'bg_color': 'gray', 'font_color': 'white', 'align':'center'}) #Formatação da celula
    planilha.set_column('A1:Q1', 25) #Formatação da coluna
    planilha.write('A1', 'Máquina', cell_format)#Cabeçalho
    planilha.write('B1', 'Nome da rotina', cell_format)
    planilha.write('C1', 'QTD Placas', cell_format)
    planilha.write('D1', 'Data de início', cell_format)
    planilha.write('E1', 'Data de término', cell_format)
    planilha.write('F1', 'Hora de início Cadastro da Rotina', cell_format)
    planilha.write('G1', 'Hora de Início (Start da Rotina)', cell_format)
    planilha.write('H1', 'Hora do término', cell_format)
    planilha.write('I1', 'Dia da Semana', cell_format)
    planilha.write('J1', 'Turno', cell_format)
    planilha.write('K1', 'QTD Resultados Liberados', cell_format)
    planilha.write('L1', 'Qtd Amostras', cell_format)
    planilha.write('M1', 'Placas processadas', cell_format)
    planilha.write('N1', 'Tempo de Carregamento da Rotina', cell_format)
    planilha.write('O1', 'Tempo Processamento Placas', cell_format)
    planilha.write('P1', 'Tempo Total Rotina (Montagem + Execução)', cell_format)
    planilha.write('Q1', 'Intervalo entre rotinas', cell_format)
    planilha.autofilter('A1:Q1')
    arquivo.close() #Fechando o Excel
#-------------------------------------------------------------------------------------------------------------    
arqs_pdf = os.listdir(diretorio + '\PDF') #Entrando no diretório dos PDF e lendo os arquivos
samples_file = diretorio + '\PDF\Samples.pdf'#Ler o Samples
def samples(path):
    pdf_file = PdfReader(open(path,'rb'))
    tabula_file = tabula.convert_into(path, 'Samples.csv', output_format='csv', pages='all')
    df = pd.read_csv(diretorio + '\Excel\Samples.csv', encoding = 'ISO-8859-1', on_bad_lines='skip')
    df.head()
samples(samples_file)
target_final = "Target_Layout.pdf" #Definindo a variavel
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
for i in arqs_pdf: #Selecionando o arquivo Target
    if i.endswith(target_final):
        target_name = diretorio_pdf + '\\' + i#Ler o PDF do Target
        target_file = PdfReader(open(target_name,'rb'))#Transformando o Target em texto
        primeira_pagina = target_file.pages[0]#Transformando a primeira página
        primeira_content = primeira_pagina.extract_text()#extrai apenas o texto
        primeira_parsed = ''.join(primeira_content)#faz a junção das linhas
        texto_primeirapagina = primeira_parsed      
        target_content = []#Target Completo
        count = len(target_file.pages)
        for i in range(count):
            target_page = target_file.pages[i]
            target_content += target_page.extract_text()#extrai apenas o texto
            texto_target = ''.join(target_content)#faz a junção das linhas
#----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        nome = "ELW 1"#Nome da maquina
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        worklist = (re.findall(r"Worklist: \s+(\w+)", texto_primeirapagina)).pop(0)#Pegando a worklist
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        ID = len(re.findall(r"ID: \d{2}", texto_target))#Quantidade de Placas
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        data = (re.findall(r'Date: \s+(\d+\/\d+\/\d+)', texto_primeirapagina)).pop(0)#Data de inicio
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        os.chdir(diretorio + '\PDF')#ACHAR O RESULT REPORT:
        rr_list = os.listdir()
        for r in rr_list:
            if r.startswith(worklist) and fnmatch.fnmatch(r, '*_[0-9][0-9]_*'): 
                rr_name = diretorio_pdf + "\\" + r
                resultreport_file = PdfReader(open(rr_name, 'rb'))
                report_page = resultreport_file.pages[0]#Transformando a primeira página em texto
                report_content = report_page.extract_text()#extrai apenas o texto
                report_parsed = ''.join(report_content)#faz a junção das linhas 
                texto_report = report_parsed
                data_finish = (re.findall(r'\s+(\d+\/\d+\/\d+)', texto_report)).pop(0)#Pegando a Data de termino
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
                hora_finish = ((re.findall(r'\s+(\d+\:\d+\s+\w+)', texto_report))).pop(0)#Pegando a Hora de término
                hora_finish = datetime.strptime(hora_finish, "%I:%M %p")
                hora_finish = datetime.strftime(hora_finish, "%H:%M")
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        hora_inicio = (re.findall(r'_\d{4}', texto_primeirapagina)).pop(0)#Pegar a Hora de inicio 
        hora_inicio = (re.findall(r'\d{4}', hora_inicio)).pop(0)
        hora_inicio = str(hora_inicio)[:2]+':'+str(hora_inicio)[2:]
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        hora_start = ((re.findall(r'\s+(\d+\:\d+\s+\w+)', texto_primeirapagina))).pop(0)#Pegando a Hora de inicio (Start da rotina)
        hora_start = datetime.strptime(hora_start, "%I:%M %p")
        hora_start = datetime.strftime(hora_start, "%H:%M")
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        def semana(data):
            dia_da_semana = datetime.strptime(data, '%m/%d/%Y')
            dia_da_semana = datetime.strftime(dia_da_semana, "%A")
            if dia_da_semana == "Monday":
                dia_da_semana = "Seg"
            elif dia_da_semana == "Tuesday":
                dia_da_semana = "Ter"
            elif dia_da_semana == "Wednesday":
                dia_da_semana = "Qua"
            elif dia_da_semana == "Thursday":
                dia_da_semana = "Qui"
            elif dia_da_semana == "Friday":
                dia_da_semana = "Sex"
            elif dia_da_semana == "Saturday":
                dia_da_semana = "Sáb"
            elif dia_da_semana == "Sunday":
                dia_da_semana = "Dom"
            return dia_da_semana
        dds = semana(data)#Pegando o Dia da semana
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------        t = turno(hora_inicio)#Pegando o turno
        def turno(hora_turno):
            hora1 = int(300)
            hora2 = int(1080)
            h1 = hora_turno[:2]
            h2 = hora_turno[2:]
            h2 = (re.findall(r'\d{2}', h2)).pop(0)
            A = [h1]
            B = [h2]
            res = [ele.lstrip('0') for ele in A].pop(0)
            res2 = [ele.lstrip('0') for ele in B].pop(0)
            h1int = int(res)
            h2int = int(res2)
            horas_minutos = h1int * 60
            minutos = int(horas_minutos + h2int)
            d = "Diurno"
            n = "Noturno" 
            if (minutos >= hora1):
                return d
            elif (minutos >= hora2):
                return n
            else:
                return n      
        t = turno(hora_inicio)#Pegando o turno
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------        t = turno(hora_inicio)#Pegando o turno
        qtd_rl = len(re.findall(r'\d{10}', texto_target)) #Pegando QTD Resultados Liberados
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------        t = turno(hora_inicio)#Pegando o turno
        palavra = worklist#Pegando Qtd Amostras
        with open(diretorio + '\Excel\Samples.csv') as f:
            qtd_a = f.read().count(palavra)
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        placas = str(round(qtd_rl/96, 2)).replace('.',',')#Pegando Placas processadas
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        temp_start = datetime.strptime(hora_start, "%H:%M")#Pegando Tempo de Carregamento da Rotina
        temp_start = temp_start - datetime(1900, 1, 1)
        temp_inicio = datetime.strptime(hora_inicio, "%H:%M")
        temp_inicio = temp_inicio - datetime(1900, 1, 1)
        TCR = temp_start - temp_inicio
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        data_i = datetime.strptime(data, '%m/%d/%Y')#Tempo Processamento Placas
        hora_is = datetime.strptime(hora_start, "%H:%M") - datetime(1900, 1, 1)
        data_f = datetime.strptime(data_finish, '%m/%d/%Y')
        hora_f = datetime.strptime(hora_finish, "%H:%M") - datetime(1900, 1, 1)
        TPP = (data_f + hora_f) - (data_i + hora_is)
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        TTP = TCR + TPP#Tempo Total Rotina (Montagem + Execução)
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        def excel(excel, dados, dados2, dados3, dados4, dados5, dados6, dados7, dados8, dados9, dados10, dados11, dados12, dados13, dados14, dados15, dados16):
            elw = load_workbook(excel)
            elw1 = elw.active
            celula = 2
            for i in elw1.iter_rows(min_row=1):
                for cell in i:
                    if cell.value == "ELW 1":
                        celula += 1
                        elw1[f"A{celula}"].value = dados
                        elw1[f'B{celula}'].value = dados2
                        elw1[f'C{celula}'].value = dados3
                        elw1[f'D{celula}'].value = dados4
                        elw1[f'E{celula}'].value = dados5
                        elw1[f'F{celula}'].value = dados6
                        elw1[f'G{celula}'].value = dados7
                        elw1[f'H{celula}'].value = dados8
                        elw1[f'I{celula}'].value = dados9
                        elw1[f'J{celula}'].value = dados10
                        elw1[f'K{celula}'].value = dados11
                        elw1[f'L{celula}'].value = dados12
                        elw1[f'M{celula}'].value = dados13
                        elw1[f'N{celula}'].value = dados14
                        elw1[f'O{celula}'].value = dados15
                        elw1[f'P{celula}'].value = dados16
                    else:
                        elw1['A2'] = dados
                        elw1['B2'] = dados2
                        elw1['C2'] = dados3
                        elw1['D2'] = dados4
                        elw1['E2'] = dados5
                        elw1['F2'] = dados6
                        elw1['G2'] = dados7
                        elw1['H2'] = dados8
                        elw1['I2'] = dados9
                        elw1['J2'] = dados10
                        elw1['K2'] = dados11
                        elw1['L2'] = dados12
                        elw1['M2'] = dados13
                        elw1['N2'] = dados14
                        elw1['O2'] = dados15
                        elw1['P2'] = dados16
            elw.save(excel)
        lista = nome, worklist, ID, data, data_finish, hora_inicio, hora_start, hora_finish, dds, t, qtd_rl, qtd_a, placas, TCR, TPP, TTP
        excel_nome = diretorio + r'\Excel\ELW ELISA 1 -  Pardini - Produtividade 14 Dias.xlsx'
        excel(excel_nome, nome, worklist, ID, data, data_finish, hora_inicio, hora_start, hora_finish, dds, t, qtd_rl, qtd_a, placas, TCR, TPP, TTP)
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
csv_file = diretorio + r'\Excel\Samples.csv'
os.remove(csv_file)
#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
print ('Tarefa completa')
